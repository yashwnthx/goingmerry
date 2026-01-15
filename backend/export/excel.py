"""Excel Export"""
from datetime import datetime, timezone
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    HEADER_FONT = Font(bold=True, size=11)
    HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
    CELL_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
    BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    def export(self, doc_model: dict) -> bytes:
        if doc_model.get("type") != "excel":
            raise ValueError(f"Cannot export '{doc_model.get('type')}' as Excel")
        
        wb = Workbook()
        wb.remove(wb.active)
        self._set_properties(wb, doc_model)
        
        sheets = doc_model.get("sheets", [])
        if not sheets:
            raise ValueError("Excel document must have at least one sheet")
        
        for sheet in sheets:
            self._add_sheet(wb, sheet)
        
        return self._to_bytes(wb)
    
    def _set_properties(self, wb: Workbook, doc_model: dict) -> None:
        wb.properties.creator = "Merry"
        wb.properties.lastModifiedBy = "Merry"
        wb.properties.title = doc_model.get("title", "Untitled")
        
        meta = doc_model.get("meta", {})
        if meta.get("created_at"):
            try:
                wb.properties.created = datetime.fromisoformat(meta["created_at"].replace("Z", "+00:00"))
            except (ValueError, TypeError):
                wb.properties.created = datetime(2024, 1, 1, tzinfo=timezone.utc)
    
    def _add_sheet(self, wb: Workbook, sheet_data: dict) -> None:
        name = sheet_data.get("name", "Sheet")[:31]
        ws = wb.create_sheet(title=name)
        
        columns = sheet_data.get("columns", [])
        rows = sheet_data.get("rows", [])
        col_map = {}
        
        for idx, col in enumerate(columns, 1):
            col_id = col.get("id")
            col_name = col.get("name", f"Column {idx}")
            col_map[col_id] = idx
            
            cell = ws.cell(row=1, column=idx, value=col_name)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGN
            cell.border = self.BORDER
        
        for row_idx, row in enumerate(rows, 2):
            cells = row.get("cells", {})
            for col_id, col_idx in col_map.items():
                cell = ws.cell(row=row_idx, column=col_idx, value=cells.get(col_id))
                cell.alignment = self.CELL_ALIGN
                cell.border = self.BORDER
        
        for idx, col in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(idx)].width = max(10, min(50, len(col.get("name", "")) + 5))
        
        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    
    def _to_bytes(self, wb: Workbook) -> bytes:
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()


def export_excel_document(doc_model: dict) -> bytes:
    return ExcelExporter().export(doc_model)
